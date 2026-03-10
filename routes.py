from flask import render_template, request, redirect, url_for, flash, jsonify, session, send_file
from app import app, db
from models import Prediction, DiscoveredPlayer, PredictionFeedback, PageView, User, BugReport, SupportChat
from flask_login import current_user
from replit_auth import make_replit_blueprint

app.register_blueprint(make_replit_blueprint(), url_prefix="/auth")

@app.before_request
def make_session_permanent():
    """Ensure all browser sessions persist across requests."""
    session.permanent = True

import io
import re
from prediction import (
    IPL_VENUES, ALL_VENUES, get_venues_by_country, Batsman, Bowler, MatchState, predict_score, validate_bowling_rules, FORMAT_CONFIG, get_venue_timezone, FIELDING_TEAMS
)
from insights import build_match_insights
from prematch import build_prematch_analysis
from player_data import (
    get_batsman_stats, get_bowler_stats, get_player_form_rating,
    PLAYER_CAREER_STATS, BOWLER_CAREER_STATS
)
from t20_players import get_all_t20_players, get_t20_batsmen, get_t20_bowlers, search_t20_players, get_all_players_including_discovered
from stats_scraper import scrape_player_t20_stats, get_cached_stats, PLAYER_ID_MAP, search_espncricinfo_players, get_or_discover_player
from live_match_scraper import get_cached_live_matches, get_match_details, refresh_live_matches

ALL_CRICKET_NATIONS = [
    "Afghanistan", "Argentina", "Australia", "Bahrain", "Bangladesh", "Bermuda",
    "Botswana", "Brazil", "Cambodia", "Cameroon", "Canada", "Chile", "China",
    "Costa Rica", "Croatia", "Czech Republic", "Denmark", "England", "Fiji",
    "Finland", "France", "Germany", "Ghana", "Gibraltar", "Greece", "Guernsey",
    "Hong Kong", "Hungary", "India", "Indonesia", "Iran", "Ireland", "Israel",
    "Italy", "Japan", "Jersey", "Kenya", "Kuwait", "Malaysia", "Maldives",
    "Mali", "Malta", "Mexico", "Mozambique", "Myanmar", "Namibia", "Nepal",
    "Netherlands", "New Zealand", "Nigeria", "Norway", "Oman",
    "Pakistan", "Panama", "Papua New Guinea", "Peru", "Philippines", "Portugal",
    "Qatar", "Romania", "Rwanda", "Samoa", "Saudi Arabia", "Scotland",
    "Sierra Leone", "Singapore", "South Africa", "South Korea", "Spain",
    "Sri Lanka", "Suriname", "Sweden", "Switzerland", "Tanzania", "Thailand",
    "Tonga", "Turkey", "UAE", "USA", "Uganda", "Vanuatu",
    "West Indies", "Zimbabwe",
    "Cayman Islands", "Eswatini", "Isle of Man", "Lesotho",
]
ALL_CRICKET_TEAMS = sorted(
    set(ALL_CRICKET_NATIONS + [n + " W" for n in ALL_CRICKET_NATIONS] + list(FIELDING_TEAMS))
)

ADDITIONAL_VENUE_COUNTRIES = {
    "Bayuemas Oval, Kuala Lumpur": "Malaysia",
    "Kinrara Academy Oval, Kuala Lumpur": "Malaysia",
    "Royal Selangor Club, Kuala Lumpur": "Malaysia",
    "UKM-YSD Cricket Oval, Bangi": "Malaysia",
    "Johor Cricket Academy Ground, Johor Bahru": "Malaysia",
    "Church Road Ground, Kuala Lumpur": "Malaysia",
    "National Cricket Ground, Bahrain": "Bahrain",
    "Al Fateh Cricket Ground, Bahrain": "Bahrain",
    "Sheikh Kamal International Cricket Stadium, Cox's Bazar": "Bangladesh",
    "Sher-e-Bangla Stadium, Mymensingh": "Bangladesh",
    "Fatullah Khan Saheb Osman Ali Stadium, Fatullah": "Bangladesh",
    "Khan Shaheb Osman Ali Stadium, Fatullah": "Bangladesh",
    "Sheikh Abu Naser Stadium, Khulna": "Bangladesh",
    "MA Aziz Stadium, Chittagong": "Bangladesh",
    "National Stadium, St George's, Grenada": "West Indies",
    "Daren Sammy National Cricket Stadium, Gros Islet": "West Indies",
    "Windsor Park, Roseau": "West Indies",
    "Arnos Vale Ground, Kingstown": "West Indies",
    "National Cricket Stadium, St Peter": "West Indies",
    "Coolidge Cricket Ground, Antigua": "West Indies",
    "Conaree Cricket Centre, Basseterre": "West Indies",
    "Jimmy Adams Oval, Kingston": "West Indies",
    "Arnos Vale Cricket Ground, St Vincent": "West Indies",
    "Recreation Ground, St John's, Antigua": "West Indies",
    "Sir Frank Worrell Memorial Ground, Bridgetown": "West Indies",
    "Beausejour Stadium, Gros Islet": "West Indies",
    "Guyana National Stadium, Georgetown": "West Indies",
    "Diego Martin Sporting Complex, Trinidad": "West Indies",
    "Central Broward Regional Park, Lauderhill": "USA",
    "Moosa Stadium, Pearland": "USA",
    "Prairie View Cricket Complex, Houston": "USA",
    "Indianapolis World Sports Park, Indianapolis": "USA",
    "Church Street Park, Morrisville": "USA",
    "Senwes Park, Potchefstroom": "South Africa",
    "Buffalo Park, East London": "South Africa",
    "Mangaung Oval, Bloemfontein": "South Africa",
    "Diamond Oval, Kimberley": "South Africa",
    "De Beers Diamond Oval, Kimberley": "South Africa",
    "Willowmoore Park, Benoni": "South Africa",
    "Willowmoore Park Cricket Ground, Benoni": "South Africa",
    "Affies Park, Pretoria": "South Africa",
    "Pietermaritzburg Oval, Pietermaritzburg": "South Africa",
    "City Oval, Pietermaritzburg": "South Africa",
    "New Wanderers Stadium, Johannesburg": "South Africa",
    "Boland Bank Park, Paarl": "South Africa",
    "Centurion Park, Centurion": "South Africa",
    "Chevrolet Park, Bloemfontein": "South Africa",
    "Sahara Stadium Kingsmead, Durban": "South Africa",
    "Old Hararians, Harare": "Zimbabwe",
    "Takashinga Sports Club, Highfield": "Zimbabwe",
    "Bulawayo Athletic Club, Bulawayo": "Zimbabwe",
    "ICC Academy, Dubai": "UAE",
    "Tolerance Oval, Abu Dhabi": "UAE",
    "Dubai Sports City Cricket Stadium, Dubai": "UAE",
    "Oman Cricket Ground, Al Amerat": "Oman",
    "AI Amerat Cricket Ground Oman Cricket, Al Amerat": "Oman",
    "Amerat North 2, Al Amerat": "Oman",
    "West End Park International Cricket Stadium, Doha": "Qatar",
    "Asian Town International Cricket Stadium, Doha": "Qatar",
    "King Fahd Security College Ground, Riyadh": "Saudi Arabia",
    "Kuwait International Cricket Ground, Kuwait City": "Kuwait",
    "TU Cricket Ground, Kirtipur": "Nepal",
    "Mulpani Cricket Ground, Kathmandu": "Nepal",
    "CAG Isipathana Ground, Colombo": "Sri Lanka",
    "P Sara Oval, Colombo": "Sri Lanka",
    "Rangiri Dambulla International Stadium, Dambulla": "Sri Lanka",
    "Suriyawewa Mahinda Rajapaksa International Cricket Stadium, Hambantota": "Sri Lanka",
    "Sugathadasa Stadium, Colombo": "Sri Lanka",
    "Colts Cricket Club Ground, Colombo": "Sri Lanka",
    "Nondescripts Cricket Club Ground, Colombo": "Sri Lanka",
    "Colts Ground, Colombo": "Sri Lanka",
    "Welagedara Stadium, Kurunegala": "Sri Lanka",
    "Bloomfield Cricket and Athletic Club, Colombo": "Sri Lanka",
    "Asgiriya Stadium, Kandy": "Sri Lanka",
    "P. Saravanamuttu Stadium, Colombo": "Sri Lanka",
    "De Saram Oval, Colombo": "Sri Lanka",
    "Rangiri Dambulla Stadium, Dambulla": "Sri Lanka",
    "Thiruvananthapuram International Cricket Stadium, Thiruvananthapuram": "India",
    "Greenfield International Stadium, Thiruvananthapuram": "India",
    "Barsapara Cricket Stadium, Guwahati": "India",
    "ACA-VDCA Cricket Stadium, Visakhapatnam": "India",
    "Vidarbha Cricket Association Stadium, Nagpur": "India",
    "Punjab Cricket Association IS Bindra Stadium, Mohali": "India",
    "Shaheed Veer Narayan Singh International Stadium, Raipur": "India",
    "BRSABV Ekana Cricket Stadium, Lucknow": "India",
    "Green Park Stadium, Kanpur": "India",
    "Nehru Stadium, Kochi": "India",
    "Dr. Y.S. Rajasekhara Reddy ACA-VDCA Cricket Stadium, Visakhapatnam": "India",
    "Birsa Munda International Cricket Stadium, Ranchi": "India",
    "Bharat Ratna Shri Atal Bihari Vajpayee Ekana Cricket Stadium, Lucknow": "India",
    "Lalabhai Contractor Stadium, Surat": "India",
    "Jawaharlal Nehru Stadium, Delhi": "India",
    "Feroz Shah Kotla, Delhi": "India",
    "Greater Noida Sports Complex Ground, Greater Noida": "India",
    "Air Force Complex Ground, Palam": "India",
    "Nahar Singh International Cricket Stadium, Faridabad": "India",
    "Sawai Man Singh Stadium, Jaipur": "India",
    "Rajasthan Cricket Association Stadium, Jaipur": "India",
    "Maharani Usharaje Trust Cricket Ground, Indore": "India",
    "Captain Roop Singh Stadium, Gwalior": "India",
    "Emerald High School Ground, Indore": "India",
    "Sector 16 Stadium, Chandigarh": "India",
    "IS Bindra Stadium, Mohali": "India",
    "Gandhi Sports Complex Ground, Amritsar": "India",
    "Reliance Stadium, Vadodara": "India",
    "Moti Bagh Stadium, Baroda": "India",
    "Sardar Patel Gujarat Stadium, Ahmedabad": "India",
    "Brabourne Stadium CCI, Mumbai": "India",
    "KSCA Cricket Ground, Hubli": "India",
    "KSCA Cricket Ground, Mysore": "India",
    "KL Saini Ground, Jaipur": "India",
    "Lalabhai Contractor Stadium, Baroda": "India",
    "Subrata Roy Sahara Stadium, Pune": "India",
    "Corporation Stadium, Kochi": "India",
    "Jawaharlal Nehru International Stadium, Kochi": "India",
    "Amingaon Cricket Stadium, Guwahati": "India",
    "CAB Academy Ground, Kalyani": "India",
    "Lal Bahadur Shastri Cricket Stadium, Hyderabad": "India",
    "Sector 16 Cricket Stadium, Chandigarh": "India",
    "Gymkhana Club Ground, Nairobi": "Kenya",
    "Ruaraka Sports Club, Nairobi": "Kenya",
    "Mombasa Sports Club, Mombasa": "Kenya",
    "Fantasy Cricket Ground, Mombasa": "Kenya",
    "Lugogo Cricket Ground, Kampala": "Uganda",
    "Kyambogo Cricket Oval, Kampala": "Uganda",
    "Entebbe Cricket Oval, Entebbe": "Uganda",
    "Tylers Standard Cricket Ground, Kampala": "Uganda",
    "Kigali Cricket Stadium, Kigali": "Rwanda",
    "Riverside Ground, Chester-le-Street": "England",
    "County Ground, Derby": "England",
    "County Ground, Hove": "England",
    "County Ground, Northampton": "England",
    "County Ground, Taunton": "England",
    "County Ground, Chelmsford": "England",
    "Kennington Oval, London": "England",
    "Grace Road, Leicester": "England",
    "New Road, Worcester": "England",
    "Wantage Road, Northampton": "England",
    "Seat Unique Riverside, Chester-le-Street": "England",
    "Uptonsteel County Ground, Leicester": "England",
    "Utilita Bowl, Southampton": "England",
    "SWALEC Stadium, Cardiff": "Wales",
    "Grange Cricket Club, Edinburgh": "Scotland",
    "Forthill, Dundee": "Scotland",
    "Stormont, Belfast": "Ireland",
    "Civil Service Cricket Club, Belfast": "Ireland",
    "Clontarf Cricket Club Ground, Dublin": "Ireland",
    "Castle Avenue, Dublin": "Ireland",
    "Civil Service Cricket Club, Stormont": "Ireland",
    "Allan Border Field, Brisbane": "Australia",
    "Junction Oval, Melbourne": "Australia",
    "North Sydney Oval, Sydney": "Australia",
    "Cazaly's Stadium, Cairns": "Australia",
    "Traeger Park, Alice Springs": "Australia",
    "Blacktown International Sportspark, Sydney": "Australia",
    "Karen Rolton Oval, Adelaide": "Australia",
    "Great Barrier Reef Arena, Mackay": "Australia",
    "University Oval, Dunedin": "New Zealand",
    "Pukekura Park, New Plymouth": "New Zealand",
    "Saxton Oval, Nelson": "New Zealand",
    "FMG Stadium Waikato, Hamilton": "New Zealand",
    "Lincoln Green, Lincoln": "New Zealand",
    "John Davies Oval, Queenstown": "New Zealand",
    "Bert Sutcliffe Oval, Lincoln": "New Zealand",
    "Sportpark Westvliet, The Hague": "Netherlands",
    "Sportpark Maarschalkerweerd, Utrecht": "Netherlands",
    "Hazelaarweg, Rotterdam": "Netherlands",
    "Sportpark Het Schootsveld, Deventer": "Netherlands",
    "Pierre Werner Cricket Ground, Walferdange": "Luxembourg",
    "Bayer Uerdingen Cricket Ground, Krefeld": "Germany",
    "Zuoz Cricket Ground, Zuoz": "Switzerland",
    "Marsa Sports Club, Marsa": "Malta",
    "Terdthai Cricket Ground, Bangkok": "Thailand",
    "Asian Institute of Technology Ground, Bangkok": "Thailand",
    "Sano International Cricket Ground, Sano": "Japan",
    "Royal Complex Cricket Ground, Sano": "Japan",
    "Yeonhui Cricket Ground, Incheon": "South Korea",
    "Moara Vlasiei Cricket Ground, Bucharest": "Romania",
    "King George V Sports Ground, Casablanca": "Morocco",
    "Grainville, Jersey": "Jersey",
    "Les Quennevais, Jersey": "Jersey",
    "KGV Playing Fields, Guernsey": "Guernsey",
    "Mission Road Ground, Bermuda": "Bermuda",
    "National Stadium, Hamilton, Bermuda": "Bermuda",
    "Faleata Oval, Apia": "Samoa",
    "Independence Park, Port Vila": "Vanuatu",
    "Independencia Cricket Ground, Asuncion": "Paraguay",
    "Maple Leaf North-West Ground, King City": "Canada",
    "Minor Ball Cricket Ground, King City": "Canada",
    "CAA Centre, Brampton": "Canada",
    "Rawalpindi Stadium, Rawalpindi": "Pakistan",
    "Iqbal Stadium, Faisalabad": "Pakistan",
    "Niaz Stadium, Hyderabad": "Pakistan",
    "Bugti Stadium, Quetta": "Pakistan",
    "Pindi Club Ground, Rawalpindi": "Pakistan",
    "Sheikhupura Stadium, Sheikhupura": "Pakistan",
    "SportsHub Oval, Singapore": "Singapore",
    "Indian Association Ground, Singapore": "Singapore",
    "SG Pipers Ground, Singapore": "Singapore",
    "Siem Reap Cricket Ground, Siem Reap": "Cambodia",
}

EXTENDED_VENUES = sorted(set(ALL_VENUES + list(ADDITIONAL_VENUE_COUNTRIES.keys())))

def get_extended_venues_by_country():
    """Merge base prediction.py venues with ADDITIONAL_VENUE_COUNTRIES, grouped by country."""
    vbc = get_venues_by_country()
    for venue, country in ADDITIONAL_VENUE_COUNTRIES.items():
        if venue not in set(v for vlist in vbc.values() for v in vlist):
            if country not in vbc:
                vbc[country] = []
            vbc[country].append(venue)
    for country in vbc:
        vbc[country] = sorted(set(vbc[country]))
    return vbc

@app.route("/", methods=["GET"])
def index():
    """Render the main prediction form with venue, player, and feedback data."""
    all_players = get_all_players_including_discovered()
    countries = list(set(p.get("country", "Unknown") for p in all_players))
    countries.sort()
    form_data = session.pop('form_data', {})
    venues_by_country = get_extended_venues_by_country()
    positive_feedback = PredictionFeedback.query.filter_by(is_positive=True).order_by(PredictionFeedback.created_at.desc()).limit(5).all()
    return render_template("index.html", venues=EXTENDED_VENUES, venues_by_country=venues_by_country, batsmen=all_players, bowlers=all_players, all_players=all_players, countries=countries, form_data=form_data, positive_feedback=positive_feedback, fielding_teams=ALL_CRICKET_TEAMS)

@app.route("/api/players", methods=["GET"])
def api_players():
    """Search players by name, role, and gender; returns JSON list for autocomplete."""
    query = request.args.get("q", "").lower()
    role = request.args.get("role", "all")
    gender = request.args.get("gender", None)
    include_discovered = request.args.get("include_discovered", "true").lower() == "true"
    
    if include_discovered:
        base_players = get_all_players_including_discovered(gender=gender)
    else:
        base_players = get_all_t20_players(gender=gender)
    
    if role == "batsman":
        players = [p for p in base_players if "Batsman" in p.get("role", "") or "All-rounder" in p.get("role", "")]
    elif role == "bowler":
        players = [p for p in base_players if "Bowler" in p.get("role", "") or "All-rounder" in p.get("role", "")]
    else:
        players = base_players
    
    import re
    def normalize(s):
        """Strip punctuation and whitespace for fuzzy player name matching."""
        return re.sub(r"['\-\.\s]+", "", s.lower())
    
    local_results = []
    if query:
        nq = normalize(query)
        local_results = [p for p in players if nq in normalize(p["name"]) or query in p.get("country", "").lower()]
    else:
        local_results = players
    
    formatted_results = [{
        "id": p["name"],
        "text": f"{p['name']} ({p.get('country', 'Unknown')})",
        "name": p["name"],
        "country": p.get("country", "Unknown"),
        "role": p.get("role", "Player"),
        "source": "local" if not p.get("espn_id") else "discovered"
    } for p in local_results[:100]]
    
    return jsonify(formatted_results)

@app.route("/api/player/<name>", methods=["GET"])
def api_player_stats(name):
    """Return batting, bowling, and form stats for a single player."""
    bat_stats = get_batsman_stats(name)
    bowl_stats = get_bowler_stats(name)
    form_rating = get_player_form_rating(name)
    
    live_stats = get_cached_stats(name)
    
    return jsonify({
        "name": name,
        "batting": bat_stats,
        "bowling": bowl_stats,
        "form_rating": round(form_rating, 1),
        "live_stats": live_stats
    })

@app.route("/api/refresh-stats/<name>", methods=["POST"])
def refresh_player_stats(name):
    """Re-scrape a player's T20 stats from ESPNcricinfo."""
    if name not in PLAYER_ID_MAP:
        discovered = get_or_discover_player(name)
        if not discovered:
            return jsonify({"error": "Player not found on ESPNcricinfo", "success": False}), 404
    
    stats = scrape_player_t20_stats(name)
    if stats:
        return jsonify({"success": True, "stats": stats})
    return jsonify({"error": "Failed to fetch stats", "success": False}), 500

@app.route("/api/cricket-api/status", methods=["GET"])
def api_cricket_status():
    """Check whether the Cricket API key is configured."""
    try:
        from cricket_api import is_api_configured
        return jsonify({
            "configured": is_api_configured(),
            "message": "Cricket API is configured" if is_api_configured() else "Cricket API key not set"
        })
    except Exception as e:
        return jsonify({"configured": False, "error": str(e)})

@app.route("/api/cricket-api/search", methods=["GET"])
def api_cricket_search():
    """Search for players via the Cricket API; requires API key."""
    query = request.args.get("q", "")
    if not query or len(query) < 2:
        return jsonify([])
    
    try:
        from cricket_api import search_players, is_api_configured
        if not is_api_configured():
            return jsonify({"error": "Cricket API key not configured"}), 400
        
        players = search_players(query, limit=20)
        return jsonify(players)
    except Exception as e:
        app.logger.error(f"Cricket API search error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/search-cricinfo", methods=["GET"])
def api_search_cricinfo():
    """Search ESPNcricinfo directly for players"""
    query = request.args.get("q", "")
    limit = min(int(request.args.get("limit", 20)), 50)
    
    if not query or len(query) < 2:
        return jsonify({"error": "Query must be at least 2 characters", "players": []}), 400
    
    try:
        players = search_espncricinfo_players(query, limit=limit)
        return jsonify({
            "success": True,
            "query": query,
            "players": players,
            "count": len(players)
        })
    except Exception as e:
        app.logger.error(f"ESPNcricinfo search error: {e}")
        return jsonify({"error": str(e), "players": []}), 500

@app.route("/api/save-discovered-player", methods=["POST"])
def api_save_discovered_player():
    """Save a discovered player to the database for future use"""
    try:
        data = request.get_json()
        if not data or not data.get("name") or not data.get("espn_id"):
            return jsonify({"error": "Missing required fields: name and espn_id", "success": False}), 400
        
        existing = DiscoveredPlayer.query.filter_by(espn_id=data["espn_id"]).first()
        if existing:
            existing.name = data.get("name", existing.name)
            existing.country = data.get("country", existing.country)
            existing.role = data.get("role", existing.role)
            existing.batting_style = data.get("batting_style", existing.batting_style)
            existing.bowling_style = data.get("bowling_style", existing.bowling_style)
            db.session.commit()
            return jsonify({"success": True, "message": "Player updated", "player": existing.to_dict()})
        
        new_player = DiscoveredPlayer(
            name=data["name"],
            espn_id=data["espn_id"],
            country=data.get("country", "Unknown"),
            role=data.get("role", "Player"),
            batting_style=data.get("batting_style"),
            bowling_style=data.get("bowling_style")
        )
        db.session.add(new_player)
        db.session.commit()
        
        PLAYER_ID_MAP[data["name"]] = data["espn_id"]
        
        return jsonify({"success": True, "message": "Player saved", "player": new_player.to_dict()})
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error saving discovered player: {e}")
        return jsonify({"error": str(e), "success": False}), 500

@app.route("/api/live-matches", methods=["GET"])
def api_live_matches():
    """Get list of live cricket matches filtered by format"""
    try:
        match_format = request.args.get("format", "mens_t20")
        all_matches = get_cached_live_matches()
        
        is_womens = match_format.startswith("womens_")
        is_odi = match_format.endswith("_odi")
        
        filtered_matches = []
        for match in all_matches:
            match_is_womens = match.get("is_womens", False)
            match_is_odi = match.get("format", "T20") == "ODI"
            
            if match_is_womens == is_womens and match_is_odi == is_odi:
                filtered_matches.append(match)
        
        return jsonify({
            "success": True,
            "matches": filtered_matches,
            "count": len(filtered_matches)
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e),
            "matches": []
        }), 500

@app.route("/api/live-matches/refresh", methods=["POST"])
def api_refresh_live_matches():
    """Force refresh of live matches list"""
    try:
        matches = refresh_live_matches()
        return jsonify({
            "success": True,
            "matches": matches,
            "count": len(matches)
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e),
            "matches": []
        }), 500

@app.route("/api/live-match/<match_id>", methods=["GET"])
def api_match_details(match_id):
    """Get detailed data for a specific live match. Returns partial data when full details unavailable."""
    try:
        match_data = get_match_details(match_id)
        if match_data:
            venue_name = re.sub(r'\s+', ' ', match_data.get('venue', '')).strip()
            matched_venue = None
            if venue_name:
                vn_lower = venue_name.lower()
                vn_normalized = re.sub(r'\.(?!\s)', '. ', vn_lower)
                vn_normalized = re.sub(r'\s+', ' ', vn_normalized).strip()
                for v in ALL_VENUES:
                    vl = v.lower()
                    vl_normalized = re.sub(r'\.(?!\s)', '. ', vl)
                    vl_normalized = re.sub(r'\s+', ' ', vl_normalized).strip()
                    if vn_normalized == vl_normalized or vn_lower == vl or vn_lower in vl or vl in vn_lower:
                        matched_venue = v
                        break
                if not matched_venue:
                    vn_words = set(vn_lower.replace(',', '').replace('.', '').split())
                    vn_city = vn_lower.split(',')[-1].strip() if ',' in vn_lower else ''
                    best_score, best_v = 0, None
                    for v in ALL_VENUES:
                        vl = v.lower()
                        v_city = vl.split(',')[-1].strip() if ',' in vl else ''
                        v_words = set(vl.replace(',', '').replace('.', '').split())
                        score = len(vn_words & v_words - {'', 'the', 'of', 'at', 'in'})
                        if v_city and vn_city and v_city == vn_city:
                            score += 2
                        if score > best_score:
                            best_score = score
                            best_v = v
                    if best_score >= 2 and best_v:
                        matched_venue = best_v
            if matched_venue:
                match_data['venue'] = matched_venue
            if match_data.get('start_time_gmt') and matched_venue:
                try:
                    import pytz
                    from datetime import datetime as dt
                    gmt_str = match_data['start_time_gmt']
                    gmt_str_clean = gmt_str.replace(',', '')
                    current_year = dt.now().year
                    gmt_dt = dt.strptime(f"{gmt_str_clean} {current_year}", "%b %d %H:%M %Y")
                    gmt_dt = pytz.utc.localize(gmt_dt)
                    tz_name = get_venue_timezone(matched_venue)
                    venue_tz = pytz.timezone(tz_name)
                    local_dt = gmt_dt.astimezone(venue_tz)
                    match_data['start_time_local'] = local_dt.strftime("%Y-%m-%dT%H:%M")
                    match_data['start_time_utc'] = gmt_dt.strftime("%Y-%m-%dT%H:%M:%SZ")
                    match_data['venue_timezone'] = tz_name
                except Exception as e:
                    logger.debug(f"Could not convert start time: {e}")
            teams = match_data.get('teams', [])
            if teams and len(teams) >= 2:
                from prediction import TEAM_FIELDING_EFFICIENCY
                TEAM_ALIASES = {
                    "ind": "India", "aus": "Australia", "eng": "England", "nz": "New Zealand",
                    "sa": "South Africa", "pak": "Pakistan", "sl": "Sri Lanka", "ban": "Bangladesh",
                    "wi": "West Indies", "afg": "Afghanistan", "zim": "Zimbabwe", "ire": "Ireland",
                    "ned": "Netherlands", "sco": "Scotland", "nam": "Namibia", "usa": "USA",
                    "nep": "Nepal", "oma": "Oman", "uae": "UAE", "can": "Canada",
                    "ken": "Kenya", "uga": "Uganda", "png": "Papua New Guinea", "hk": "Hong Kong",
                    "india": "India", "australia": "Australia", "england": "England",
                    "new zealand": "New Zealand", "south africa": "South Africa",
                    "pakistan": "Pakistan", "sri lanka": "Sri Lanka", "bangladesh": "Bangladesh",
                    "west indies": "West Indies", "afghanistan": "Afghanistan",
                    "zimbabwe": "Zimbabwe", "ireland": "Ireland", "netherlands": "Netherlands",
                    "scotland": "Scotland", "namibia": "Namibia", "nepal": "Nepal",
                }
                match_title = match_data.get('title', '') or match_data.get('match_format', '') or ''
                is_womens = any(w in match_title.lower() for w in ['women', 'wmn', 'wodi', 'wt20'])
                def resolve_team(name):
                    """Map a raw team name/abbreviation to the canonical fielding efficiency key."""
                    nl = name.lower().strip()
                    nl = re.sub(r'\s*(women|wmn|w)\s*$', '', nl, flags=re.I).strip()
                    if nl in TEAM_ALIASES:
                        team = TEAM_ALIASES[nl]
                    else:
                        team = None
                        for key, val in TEAM_FIELDING_EFFICIENCY.items():
                            if not key.endswith(' W') and nl in key.lower():
                                team = val if isinstance(val, str) else key
                                break
                    if team and is_womens and f"{team} W" in TEAM_FIELDING_EFFICIENCY:
                        team = f"{team} W"
                    return team
                bowling_team_raw = None
                bowling_squad_idx = match_data.get('bowling_team_squad_idx')
                if bowling_squad_idx is not None and len(teams) > bowling_squad_idx:
                    bowling_team_raw = teams[bowling_squad_idx]
                if not bowling_team_raw and len(teams) >= 2:
                    innings = match_data.get('innings', 1)
                    bowling_team_raw = teams[1] if innings == 1 else teams[0]

                resolved = resolve_team(bowling_team_raw) if bowling_team_raw else None
                if resolved:
                    match_data['fielding_team'] = resolved
                elif bowling_team_raw:
                    match_data['fielding_team'] = bowling_team_raw
            return jsonify({
                "success": True,
                "match": match_data,
                "partial": match_data.get('partial', False),
                "loaded_fields": match_data.get('loaded_fields', [])
            })
        return jsonify({
            "success": False,
            "error": "Could not fetch match details"
        }), 404
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route("/predict", methods=["POST"])
def predict():
    """Handle prediction form submission: validate inputs, run engine, render results."""
    try:
        session['form_data'] = dict(request.form)
        
        input_mode = request.form.get("input_mode", "manual")
        venue = request.form.get("venue", "")
        current_score = int(request.form.get("current_score", 0))
        wickets_fallen = int(request.form.get("wickets_fallen", 0))
        overs_completed = float(request.form.get("overs_completed", 0))
        next_over_bowler = request.form.get("next_over_bowler", "").strip()
        last_over_bowler = request.form.get("last_over_bowler", "")
        innings = int(request.form.get("innings", 1))
        target = int(request.form.get("target", 0)) if request.form.get("target") else 0
        match_format = request.form.get("match_format", "mens_t20")
        fielding_team = request.form.get("fielding_team", "").strip()
        
        if match_format not in FORMAT_CONFIG:
            match_format = "mens_t20"
        
        format_config = FORMAT_CONFIG[match_format]
        max_overs = format_config['max_overs']
        max_bowler_overs = format_config['max_bowler_overs']
        is_odi = format_config['is_odi']
        format_name = "ODI" if is_odi else "T20"
        
        if current_score < 0:
            flash("Current score cannot be negative", "error")
            return redirect(url_for("index"))
        
        if wickets_fallen < 0:
            flash("Wickets fallen cannot be negative", "error")
            return redirect(url_for("index"))
        
        if wickets_fallen >= 10:
            flash("The innings has ended - all 10 wickets have fallen. No prediction possible.", "error")
            return redirect(url_for("index"))
        
        if overs_completed < 0:
            flash("Overs completed cannot be negative", "error")
            return redirect(url_for("index"))
        
        if overs_completed >= max_overs:
            flash(f"The innings has ended - all {max_overs} overs have been bowled. No prediction possible.", "error")
            return redirect(url_for("index"))
        
        overs_decimal = overs_completed - int(overs_completed)
        if overs_decimal >= 0.6:
            flash(f"Invalid overs format: {overs_completed}. One over has 6 balls, so valid decimals are .0 to .5 only", "error")
            return redirect(url_for("index"))
        
        if not venue:
            flash("Please select a cricket venue", "error")
            return redirect(url_for("index"))
        is_custom_venue = venue not in ALL_VENUES
        
        batsmen = []
        for i in range(1, 3):
            name = request.form.get(f"batsman{i}_name", "")
            runs = request.form.get(f"batsman{i}_runs", "0")
            balls = request.form.get(f"batsman{i}_balls", "0")
            
            runs_val = int(runs) if runs else 0
            balls_val = int(balls) if balls else 0
            
            if name:
                if runs_val < 0:
                    flash(f"Runs scored by {name} cannot be negative", "error")
                    return redirect(url_for("index"))
                
                if runs_val > current_score:
                    flash(f"Runs scored by {name} ({runs_val}) cannot exceed team score ({current_score})", "error")
                    return redirect(url_for("index"))
                
                if balls_val < 0:
                    flash(f"Balls faced by {name} cannot be negative", "error")
                    return redirect(url_for("index"))
                
                max_balls = max_overs * 6
                if balls_val > max_balls:
                    flash(f"Balls faced by {name} cannot exceed {max_balls} (maximum balls in a {format_name} innings)", "error")
                    return redirect(url_for("index"))
                
                batsmen.append(Batsman(
                    name=name,
                    runs=runs_val,
                    balls_faced=balls_val
                ))
        
        bowlers = []
        bowler_names = request.form.getlist("bowler_name[]")
        bowler_runs = request.form.getlist("bowler_runs[]")
        bowler_wickets = request.form.getlist("bowler_wickets[]")
        bowler_remaining = request.form.getlist("bowler_remaining[]")
        
        for i in range(len(bowler_names)):
            if bowler_names[i]:
                overs_remaining_val = float(bowler_remaining[i]) if i < len(bowler_remaining) and bowler_remaining[i] else max_bowler_overs
                overs_bowled = max_bowler_overs - overs_remaining_val
                runs_conceded = int(bowler_runs[i]) if i < len(bowler_runs) and bowler_runs[i] else 0
                wickets_taken = int(bowler_wickets[i]) if i < len(bowler_wickets) and bowler_wickets[i] else 0
                
                if overs_remaining_val < 0:
                    flash(f"Overs remaining for {bowler_names[i]} cannot be negative", "error")
                    return redirect(url_for("index"))
                
                if overs_remaining_val > max_bowler_overs:
                    flash(f"Overs remaining for {bowler_names[i]} cannot exceed {max_bowler_overs} ({format_name} rule)", "error")
                    return redirect(url_for("index"))
                
                overs_decimal = overs_remaining_val - int(overs_remaining_val)
                if overs_decimal >= 0.6:
                    flash(f"Invalid overs format for {bowler_names[i]}: {overs_remaining_val}. One over has 6 balls, so valid decimals are .0 to .5 only", "error")
                    return redirect(url_for("index"))
                
                if runs_conceded < 0:
                    flash(f"Runs conceded by {bowler_names[i]} cannot be negative", "error")
                    return redirect(url_for("index"))
                
                if wickets_taken < 0:
                    flash(f"Wickets taken by {bowler_names[i]} cannot be negative", "error")
                    return redirect(url_for("index"))
                
                if wickets_taken > 10:
                    flash(f"Wickets taken by {bowler_names[i]} cannot exceed 10 (maximum wickets in an innings)", "error")
                    return redirect(url_for("index"))
                
                bowlers.append(Bowler(
                    name=bowler_names[i],
                    overs_bowled=overs_bowled,
                    runs_conceded=runs_conceded,
                    wickets=wickets_taken,
                    overs_remaining=overs_remaining_val
                ))
        
        batsmen_to_come = []
        batsmen_to_come_names = request.form.getlist("batsman_to_come[]")
        batsmen_at_crease = [b.name for b in batsmen]
        
        for name in batsmen_to_come_names:
            if name:
                if name in batsmen_at_crease:
                    flash(f"{name} is already at the crease and cannot be in the batting order", "error")
                    return redirect(url_for("index"))
                if name in batsmen_to_come:
                    flash(f"{name} is selected multiple times in batsmen yet to bat", "error")
                    return redirect(url_for("index"))
                batsmen_to_come.append(name)
        
        all_batsmen_names = batsmen_at_crease + batsmen_to_come
        bowler_names_list = [b.name for b in bowlers]
        
        for bowler_name in bowler_names_list:
            if bowler_name in all_batsmen_names:
                flash(f"{bowler_name} is batting and cannot also be bowling. A player cannot bowl to themselves.", "error")
                return redirect(url_for("index"))
        
        bowler_names_in_figures = [b.name for b in bowlers if b.name]
        
        if next_over_bowler:
            if next_over_bowler in all_batsmen_names:
                flash(f"{next_over_bowler} is batting and cannot bowl the next over. A player cannot bowl to themselves.", "error")
                return redirect(url_for("index"))
            
            bowling_team_11_str = request.form.get("bowling_team_11", "")
            bowling_team_11 = [name.strip() for name in bowling_team_11_str.split(",") if name.strip()] if bowling_team_11_str else []
            
            if bowling_team_11:
                if next_over_bowler not in bowling_team_11:
                    flash(f"{next_over_bowler} is not in the bowling team's playing 11. Please select a bowler from the bowling team.", "error")
                    return redirect(url_for("index"))
            elif bowler_names_in_figures:
                if next_over_bowler not in bowler_names_in_figures:
                    flash(f"{next_over_bowler} is not in the bowling lineup. Please select them in the Bowling Figures section first.", "error")
                    return redirect(url_for("index"))
            
            next_bowler_obj = None
            for b in bowlers:
                if b.name == next_over_bowler:
                    next_bowler_obj = b
                    break
            
            if next_bowler_obj and next_bowler_obj.overs_remaining <= 0:
                flash(f"{next_over_bowler} has no overs remaining and cannot bowl the next over", "error")
                return redirect(url_for("index"))
        
        match_state = MatchState(
            current_score=current_score,
            wickets_fallen=wickets_fallen,
            overs_completed=overs_completed,
            venue=venue,
            batsmen=batsmen,
            bowlers=bowlers,
            next_over_bowler=next_over_bowler,
            last_over_bowler=last_over_bowler,
            batsmen_to_come=batsmen_to_come,
            innings=innings,
            target=target,
            match_format=match_format,
            fielding_team=fielding_team
        )
        
        validation_errors = validate_bowling_rules(match_state)
        if validation_errors:
            for error in validation_errors:
                flash(error, "error")
            return redirect(url_for("index"))
        
        result = predict_score(match_state)
        
        prediction = Prediction(
            venue=venue,
            current_score=current_score,
            wickets_fallen=wickets_fallen,
            overs_remaining=max_overs - overs_completed,
            predicted_final_score=result.predicted_final_score,
            predicted_wickets=result.predicted_wickets,
            predicted_next_over_runs=result.predicted_next_over_runs,
            next_over_bowler=next_over_bowler
        )
        db.session.add(prediction)
        db.session.commit()
        
        batsmen_stats = {}
        for batsman in batsmen:
            stats = get_batsman_stats(batsman.name)
            if stats:
                batsmen_stats[batsman.name] = stats
        
        bowler_stats = get_bowler_stats(next_over_bowler) if next_over_bowler else None
        
        stats_label = "Career"
        return render_template(
            "results.html",
            result=result,
            match_state=match_state,
            venues=IPL_VENUES,
            batsmen_stats=batsmen_stats,
            bowler_stats=bowler_stats,
            stats_label=stats_label,
            is_live_data=(input_mode == "live")
        )
        
    except Exception as e:
        app.logger.error(f"Prediction error: {str(e)}")
        flash(f"Error processing prediction: {str(e)}", "error")
        return redirect(url_for("index"))


@app.route("/demo")
def demo_prediction():
    """Run a demo prediction with a realistic pre-filled match scenario."""
    try:
        import random
        
        scenarios = [
            {
                "title": "India vs Australia - MCG T20",
                "match_format": "mens_t20",
                "venue": "Melbourne Cricket Ground, Melbourne",
                "current_score": 87,
                "wickets_fallen": 2,
                "overs_completed": 10.3,
                "innings": 1,
                "target": 0,
                "fielding_team": "Australia",
                "batsmen": [
                    Batsman(name="Suryakumar Yadav", runs=42, balls_faced=28),
                    Batsman(name="Hardik Pandya", runs=18, balls_faced=12),
                ],
                "bowlers": [
                    Bowler(name="Mitchell Starc", overs_bowled=3, runs_conceded=28, wickets=1, overs_remaining=1),
                    Bowler(name="Josh Hazlewood", overs_bowled=2, runs_conceded=15, wickets=1, overs_remaining=2),
                    Bowler(name="Adam Zampa", overs_bowled=2.3, runs_conceded=22, wickets=0, overs_remaining=1.3),
                    Bowler(name="Pat Cummins", overs_bowled=2, runs_conceded=14, wickets=0, overs_remaining=2),
                    Bowler(name="Glenn Maxwell", overs_bowled=1, runs_conceded=8, wickets=0, overs_remaining=3),
                ],
                "next_over_bowler": "Pat Cummins",
                "last_over_bowler": "Adam Zampa",
                "batsmen_to_come": ["Rinku Singh", "Ravindra Jadeja", "Axar Patel"],
            },
            {
                "title": "England vs South Africa - Lord's T20",
                "match_format": "mens_t20",
                "venue": "Lord's, London",
                "current_score": 112,
                "wickets_fallen": 3,
                "overs_completed": 13.2,
                "innings": 1,
                "target": 0,
                "fielding_team": "South Africa",
                "batsmen": [
                    Batsman(name="Jos Buttler", runs=55, balls_faced=34),
                    Batsman(name="Liam Livingstone", runs=24, balls_faced=16),
                ],
                "bowlers": [
                    Bowler(name="Kagiso Rabada", overs_bowled=3, runs_conceded=25, wickets=1, overs_remaining=1),
                    Bowler(name="Anrich Nortje", overs_bowled=3, runs_conceded=30, wickets=1, overs_remaining=1),
                    Bowler(name="Keshav Maharaj", overs_bowled=3, runs_conceded=22, wickets=0, overs_remaining=1),
                    Bowler(name="Marco Jansen", overs_bowled=2.2, runs_conceded=18, wickets=1, overs_remaining=1.4),
                    Bowler(name="Aiden Markram", overs_bowled=2, runs_conceded=17, wickets=0, overs_remaining=2),
                ],
                "next_over_bowler": "Kagiso Rabada",
                "last_over_bowler": "Marco Jansen",
                "batsmen_to_come": ["Harry Brook", "Moeen Ali", "Sam Curran"],
            },
            {
                "title": "Pakistan vs West Indies - Chase at Dubai",
                "match_format": "mens_t20",
                "venue": "Dubai International Cricket Stadium",
                "current_score": 95,
                "wickets_fallen": 1,
                "overs_completed": 11.0,
                "innings": 2,
                "target": 178,
                "fielding_team": "West Indies",
                "batsmen": [
                    Batsman(name="Babar Azam", runs=52, balls_faced=38),
                    Batsman(name="Fakhar Zaman", runs=30, balls_faced=22),
                ],
                "bowlers": [
                    Bowler(name="Alzarri Joseph", overs_bowled=3, runs_conceded=24, wickets=1, overs_remaining=1),
                    Bowler(name="Akeal Hosein", overs_bowled=3, runs_conceded=20, wickets=0, overs_remaining=1),
                    Bowler(name="Andre Russell", overs_bowled=2, runs_conceded=22, wickets=0, overs_remaining=2),
                    Bowler(name="Gudakesh Motie", overs_bowled=2, runs_conceded=16, wickets=0, overs_remaining=2),
                    Bowler(name="Romario Shepherd", overs_bowled=1, runs_conceded=13, wickets=0, overs_remaining=3),
                ],
                "next_over_bowler": "Andre Russell",
                "last_over_bowler": "Akeal Hosein",
                "batsmen_to_come": ["Mohammad Rizwan", "Shadab Khan", "Shaheen Afridi"],
            },
            {
                "title": "Australia Women vs India Women - Sydney T20",
                "match_format": "womens_t20",
                "venue": "Sydney Cricket Ground, Sydney",
                "current_score": 78,
                "wickets_fallen": 2,
                "overs_completed": 11.4,
                "innings": 1,
                "target": 0,
                "fielding_team": "India Women",
                "batsmen": [
                    Batsman(name="Beth Mooney", runs=38, balls_faced=30),
                    Batsman(name="Ellyse Perry", runs=22, balls_faced=18),
                ],
                "bowlers": [
                    Bowler(name="Deepti Sharma", overs_bowled=3, runs_conceded=18, wickets=1, overs_remaining=1),
                    Bowler(name="Renuka Singh", overs_bowled=3, runs_conceded=22, wickets=1, overs_remaining=1),
                    Bowler(name="Radha Yadav", overs_bowled=2.4, runs_conceded=15, wickets=0, overs_remaining=1.2),
                    Bowler(name="Shreyanka Patil", overs_bowled=2, runs_conceded=14, wickets=0, overs_remaining=2),
                    Bowler(name="Pooja Vastrakar", overs_bowled=1, runs_conceded=9, wickets=0, overs_remaining=3),
                ],
                "next_over_bowler": "Deepti Sharma",
                "last_over_bowler": "Radha Yadav",
                "batsmen_to_come": ["Ashleigh Gardner", "Tahlia McGrath", "Grace Harris"],
            },
            {
                "title": "India vs England - Wankhede ODI",
                "match_format": "mens_odi",
                "venue": "Wankhede Stadium, Mumbai",
                "current_score": 172,
                "wickets_fallen": 3,
                "overs_completed": 32.0,
                "innings": 1,
                "target": 0,
                "fielding_team": "England",
                "batsmen": [
                    Batsman(name="Virat Kohli", runs=68, balls_faced=74),
                    Batsman(name="KL Rahul", runs=35, balls_faced=38),
                ],
                "bowlers": [
                    Bowler(name="Jofra Archer", overs_bowled=7, runs_conceded=38, wickets=1, overs_remaining=3),
                    Bowler(name="Mark Wood", overs_bowled=6, runs_conceded=42, wickets=1, overs_remaining=4),
                    Bowler(name="Adil Rashid", overs_bowled=7, runs_conceded=32, wickets=1, overs_remaining=3),
                    Bowler(name="Liam Livingstone", overs_bowled=5, runs_conceded=28, wickets=0, overs_remaining=5),
                    Bowler(name="Chris Woakes", overs_bowled=7, runs_conceded=32, wickets=0, overs_remaining=3),
                ],
                "next_over_bowler": "Adil Rashid",
                "last_over_bowler": "Jofra Archer",
                "batsmen_to_come": ["Hardik Pandya", "Ravindra Jadeja", "Axar Patel"],
            },
            {
                "title": "England Women vs South Africa Women - Edgbaston ODI",
                "match_format": "womens_odi",
                "venue": "Edgbaston, Birmingham",
                "current_score": 145,
                "wickets_fallen": 2,
                "overs_completed": 30.0,
                "innings": 1,
                "target": 0,
                "fielding_team": "South Africa Women",
                "batsmen": [
                    Batsman(name="Nat Sciver-Brunt", runs=62, balls_faced=68),
                    Batsman(name="Amy Jones", runs=28, balls_faced=35),
                ],
                "bowlers": [
                    Bowler(name="Marizanne Kapp", overs_bowled=7, runs_conceded=30, wickets=1, overs_remaining=3),
                    Bowler(name="Shabnim Ismail", overs_bowled=6, runs_conceded=35, wickets=1, overs_remaining=4),
                    Bowler(name="Nonkululeko Mlaba", overs_bowled=7, runs_conceded=28, wickets=0, overs_remaining=3),
                    Bowler(name="Masabata Klaas", overs_bowled=5, runs_conceded=25, wickets=0, overs_remaining=5),
                    Bowler(name="Ayabonga Khaka", overs_bowled=5, runs_conceded=27, wickets=0, overs_remaining=5),
                ],
                "next_over_bowler": "Marizanne Kapp",
                "last_over_bowler": "Nonkululeko Mlaba",
                "batsmen_to_come": ["Heather Knight", "Sophie Ecclestone", "Kate Cross"],
            },
        ]
        
        scenario = random.choice(scenarios)
        
        match_state = MatchState(
            current_score=scenario["current_score"],
            wickets_fallen=scenario["wickets_fallen"],
            overs_completed=scenario["overs_completed"],
            venue=scenario["venue"],
            batsmen=scenario["batsmen"],
            bowlers=scenario["bowlers"],
            next_over_bowler=scenario["next_over_bowler"],
            last_over_bowler=scenario["last_over_bowler"],
            batsmen_to_come=scenario["batsmen_to_come"],
            innings=scenario["innings"],
            target=scenario["target"],
            match_format=scenario["match_format"],
            fielding_team=scenario["fielding_team"]
        )
        
        result = predict_score(match_state)
        
        batsmen_stats = {}
        for batsman in scenario["batsmen"]:
            stats = get_batsman_stats(batsman.name)
            if stats:
                batsmen_stats[batsman.name] = stats
        
        bowler_stats = get_bowler_stats(scenario["next_over_bowler"]) if scenario["next_over_bowler"] else None
        
        stats_label = "Career"
        return render_template(
            "results.html",
            result=result,
            match_state=match_state,
            venues=IPL_VENUES,
            batsmen_stats=batsmen_stats,
            bowler_stats=bowler_stats,
            is_demo=True,
            demo_title=scenario["title"],
            stats_label=stats_label,
            is_live_data=False
        )
        
    except Exception as e:
        app.logger.error(f"Demo prediction error: {str(e)}")
        flash(f"Error running demo: {str(e)}", "error")
        return redirect(url_for("index"))


@app.route("/feedback/submit", methods=["POST"])
def submit_feedback_form():
    """Submit user feedback on a prediction via form POST - requires login"""
    if not current_user.is_authenticated:
        flash("Please log in to submit feedback.", "warning")
        return redirect(url_for('replit_auth.login'))
    
    try:
        is_positive = request.form.get('is_positive', 'true').lower() == 'true'
        
        user_id = current_user.id
        username = current_user.display_name
        contact_email = request.form.get('contact_email', '').strip() or None
        
        if contact_email:
            import re
            if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]*[a-zA-Z][a-zA-Z0-9.-]*\.[a-zA-Z]{2,6}$', contact_email):
                flash("The email address you entered doesn't look right. Please check it and try again.", "warning")
                return redirect(url_for("index"))
        
        feedback = PredictionFeedback(
            is_positive=is_positive,
            user_id=user_id,
            username=username,
            reason_wrong_player=request.form.get('reason_wrong_player') == 'true',
            reason_unrealistic=request.form.get('reason_unrealistic') == 'true',
            reason_ui_confusing=request.form.get('reason_ui_confusing') == 'true',
            reason_bug_error=request.form.get('reason_bug_error') == 'true',
            feedback_text=request.form.get('feedback_text', '')[:300] if request.form.get('feedback_text') else None,
            contact_email=contact_email,
            venue=request.form.get('venue'),
            match_format=request.form.get('match_format'),
            current_score=int(request.form.get('current_score', 0)) if request.form.get('current_score') else None,
            predicted_score=int(request.form.get('predicted_score', 0)) if request.form.get('predicted_score') else None
        )
        
        db.session.add(feedback)
        db.session.commit()
        
        if contact_email:
            try:
                from email_responder import send_feedback_reply
                email_sent = send_feedback_reply(
                    to_email=contact_email,
                    is_positive=is_positive,
                    feedback_text=feedback.feedback_text,
                    venue=feedback.venue,
                    match_format=feedback.match_format,
                    predicted_score=feedback.predicted_score
                )
                if email_sent:
                    feedback.email_sent = True
                    db.session.commit()
            except Exception as email_err:
                app.logger.error(f"Auto-reply email error: {email_err}")
        
        flash("Thank you for your feedback! It helps us improve.", "success")
        return redirect(url_for("index"))
    except Exception as e:
        app.logger.error(f"Feedback submission error: {str(e)}")
        flash("There was an error submitting your feedback. Please try again.", "error")
        return redirect(url_for("index"))


@app.route("/feedback/export", methods=["GET"])
def export_feedback():
    """Export all feedback to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        feedback_list = PredictionFeedback.query.order_by(PredictionFeedback.created_at.desc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Prediction Feedback"
        
        headers = ["ID", "Date/Time", "Feedback Type", "Reasons", "Comments", "Venue", "Format", "Current Score", "Predicted Score"]
        header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        for row_num, feedback in enumerate(feedback_list, 2):
            data = feedback.to_dict()
            ws.cell(row=row_num, column=1, value=data['id']).border = thin_border
            ws.cell(row=row_num, column=2, value=data['created_at']).border = thin_border
            ws.cell(row=row_num, column=3, value=data['feedback_type']).border = thin_border
            ws.cell(row=row_num, column=4, value=data['reasons']).border = thin_border
            ws.cell(row=row_num, column=5, value=data['feedback_text']).border = thin_border
            ws.cell(row=row_num, column=6, value=data['venue']).border = thin_border
            ws.cell(row=row_num, column=7, value=data['match_format']).border = thin_border
            ws.cell(row=row_num, column=8, value=data['current_score']).border = thin_border
            ws.cell(row=row_num, column=9, value=data['predicted_score']).border = thin_border
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='prediction_feedback.xlsx'
        )
    except Exception as e:
        app.logger.error(f"Excel export error: {str(e)}")
        flash(f"Error exporting feedback: {str(e)}", "error")
        return redirect(url_for("index"))


@app.route("/feedback", methods=["GET"])
def view_feedback():
    """View all feedback in a simple dashboard"""
    feedback_list = PredictionFeedback.query.order_by(PredictionFeedback.created_at.desc()).limit(100).all()
    total_feedback = PredictionFeedback.query.count()
    positive_count = PredictionFeedback.query.filter_by(is_positive=True).count()
    negative_count = PredictionFeedback.query.filter_by(is_positive=False).count()
    
    return render_template(
        "feedback_dashboard.html",
        feedback_list=feedback_list,
        total_feedback=total_feedback,
        positive_count=positive_count,
        negative_count=negative_count
    )


@app.before_request
def track_page_view():
    """Track page views for analytics"""
    excluded_paths = ['/static/', '/api/', '/analytics', '/favicon']
    if any(request.path.startswith(p) for p in excluded_paths):
        return
    if request.method != 'GET':
        return
    try:
        PageView.log_view(request)
    except Exception as e:
        app.logger.error(f"Error logging page view: {e}")


@app.route("/analytics", methods=["GET"])
def analytics_dashboard():
    """View traffic analytics dashboard"""
    from datetime import datetime, timedelta
    from sqlalchemy import func, distinct
    
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_ago = now - timedelta(days=7)
    month_ago = now - timedelta(days=30)
    
    total_views = PageView.query.count()
    total_unique = db.session.query(func.count(distinct(PageView.ip_hash))).scalar() or 0
    
    today_views = PageView.query.filter(PageView.created_at >= today_start).count()
    today_unique = db.session.query(func.count(distinct(PageView.ip_hash))).filter(PageView.created_at >= today_start).scalar() or 0
    
    week_views = PageView.query.filter(PageView.created_at >= week_ago).count()
    week_unique = db.session.query(func.count(distinct(PageView.ip_hash))).filter(PageView.created_at >= week_ago).scalar() or 0
    
    month_views = PageView.query.filter(PageView.created_at >= month_ago).count()
    
    top_pages = db.session.query(
        PageView.path,
        func.count(PageView.id).label('views')
    ).group_by(PageView.path).order_by(func.count(PageView.id).desc()).limit(10).all()
    
    browser_stats = db.session.query(
        PageView.browser,
        func.count(PageView.id).label('count')
    ).group_by(PageView.browser).order_by(func.count(PageView.id).desc()).all()
    
    device_stats = db.session.query(
        PageView.device_type,
        func.count(PageView.id).label('count')
    ).group_by(PageView.device_type).order_by(func.count(PageView.id).desc()).all()
    
    daily_views = db.session.query(
        func.date(PageView.created_at).label('date'),
        func.count(PageView.id).label('views'),
        func.count(distinct(PageView.ip_hash)).label('unique')
    ).filter(PageView.created_at >= week_ago).group_by(func.date(PageView.created_at)).order_by(func.date(PageView.created_at)).all()
    
    recent_views = PageView.query.order_by(PageView.created_at.desc()).limit(20).all()
    
    return render_template(
        "analytics_dashboard.html",
        total_views=total_views,
        total_unique=total_unique,
        today_views=today_views,
        today_unique=today_unique,
        week_views=week_views,
        week_unique=week_unique,
        month_views=month_views,
        top_pages=top_pages,
        browser_stats=browser_stats,
        device_stats=device_stats,
        daily_views=daily_views,
        recent_views=recent_views
    )


@app.route("/api/venue-timezone")
def api_venue_timezone():
    """Return the IANA timezone string for a given venue name."""
    venue = request.args.get("venue", "")
    tz = get_venue_timezone(venue)
    return jsonify({"timezone": tz, "venue": venue})

@app.route("/prematch", methods=["GET", "POST"])
def prematch():
    """Render pre-match analysis page with venue conditions, par scores, and toss advice."""
    venues_by_country = get_venues_by_country()

    if request.method == "GET":
        return render_template("prematch.html",
                               venues=ALL_VENUES,
                               venues_by_country=venues_by_country,
                               analysis=None)

    venue = request.form.get("venue", "")
    match_format = request.form.get("match_format", "mens_t20")
    match_time = request.form.get("match_time", "")
    team1 = request.form.get("team1", "").strip()
    team2 = request.form.get("team2", "").strip()

    if match_format not in FORMAT_CONFIG:
        match_format = "mens_t20"

    if not venue or venue not in ALL_VENUES:
        flash("Please select a valid venue", "error")
        return redirect(url_for("prematch"))

    analysis = build_prematch_analysis(
        venue=venue,
        match_format=match_format,
        match_time=match_time if match_time else None,
        team1=team1 if team1 else None,
        team2=team2 if team2 else None
    )

    return render_template("prematch.html",
                           venues=ALL_VENUES,
                           venues_by_country=venues_by_country,
                           analysis=analysis)


@app.route("/insights", methods=["GET", "POST"])
def insights():
    """Render dismissal insights page predicting how batsmen may get out."""
    all_players = get_all_players_including_discovered()
    venues_by_country = get_venues_by_country()

    if request.method == "GET":
        return render_template("insights.html",
                               venues=ALL_VENUES,
                               venues_by_country=venues_by_country,
                               all_players=all_players,
                               insights_data=None)

    venue = request.form.get("venue", "")
    current_score = int(request.form.get("current_score", 0))
    wickets = int(request.form.get("wickets_fallen", 0))
    overs = float(request.form.get("overs_completed", 0))
    next_bowler = request.form.get("next_over_bowler", "")
    innings = int(request.form.get("innings", 1))
    target = int(request.form.get("target", 0)) if request.form.get("target") else 0
    match_format = request.form.get("match_format", "mens_t20")

    if match_format not in FORMAT_CONFIG:
        match_format = "mens_t20"

    format_config = FORMAT_CONFIG[match_format]
    max_overs = format_config['max_overs']

    if current_score < 0:
        flash("Current score cannot be negative", "error")
        return redirect(url_for("insights"))
    if wickets < 0 or wickets >= 10:
        flash("Wickets must be between 0 and 9", "error")
        return redirect(url_for("insights"))
    if overs < 0 or overs >= max_overs:
        flash(f"Overs must be between 0 and {max_overs - 1}.5", "error")
        return redirect(url_for("insights"))
    overs_decimal = overs - int(overs)
    if overs_decimal >= 0.6:
        flash(f"Invalid overs format: {overs}. Valid decimals are .0 to .5", "error")
        return redirect(url_for("insights"))
    if innings == 2 and target > 0 and current_score >= target:
        flash("Score already exceeds target", "error")
        return redirect(url_for("insights"))

    batsmen_list = []
    for i in range(1, 3):
        name = request.form.get(f"batsman{i}_name", "")
        runs = int(request.form.get(f"batsman{i}_runs", 0) or 0)
        balls = int(request.form.get(f"batsman{i}_balls", 0) or 0)
        if name:
            batsmen_list.append({'name': name, 'runs': runs, 'balls_faced': balls})

    if not batsmen_list:
        flash("Please select at least one batsman for insights", "error")
        return redirect(url_for("insights"))

    if not next_bowler:
        flash("Please select the current bowler", "error")
        return redirect(url_for("insights"))

    if not venue or venue not in ALL_VENUES:
        flash("Please select a valid venue", "error")
        return redirect(url_for("insights"))

    insights_data = build_match_insights(
        batsmen=batsmen_list,
        bowler_name=next_bowler,
        venue=venue,
        overs=overs,
        wickets=wickets,
        current_score=current_score,
        target=target,
        innings=innings,
        match_format=match_format
    )

    from prediction import get_venue_conditions
    from ipl_stats import get_venue_weather, get_pitch_tempo
    venue_conditions = get_venue_conditions(venue)
    weather_data = get_venue_weather(venue)
    pitch_tempo = get_pitch_tempo(venue)

    format_labels = {
        'mens_t20': "Men's T20",
        'womens_t20': "Women's T20",
        'mens_odi': "Men's ODI",
        'womens_odi': "Women's ODI"
    }

    return render_template("insights.html",
                           venues=ALL_VENUES,
                           venues_by_country=venues_by_country,
                           all_players=all_players,
                           insights_data=insights_data,
                           venue=venue,
                           venue_conditions=venue_conditions,
                           current_score=current_score,
                           wickets=wickets,
                           overs=overs,
                           next_bowler=next_bowler,
                           innings=innings,
                           target=target,
                           match_format=match_format,
                           format_label=format_labels.get(match_format, "T20"),
                           weather_data=weather_data,
                           pitch_tempo=pitch_tempo,
                           form_data=dict(request.form))


@app.route("/bug-report", methods=["GET", "POST"])
def bug_report():
    """Handle bug report submission with optional screenshot and email notification."""
    if request.method == "GET":
        return render_template("bug_report.html")

    import os
    import json
    import uuid

    category = request.form.get("category", "Bug")
    title = request.form.get("title", "").strip()
    description = request.form.get("description", "").strip()
    contact_email = request.form.get("contact_email", "").strip() or None
    
    if contact_email:
        import re
        if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]*[a-zA-Z][a-zA-Z0-9.-]*\.[a-zA-Z]{2,6}$', contact_email):
            is_ajax_check = (request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                       or request.accept_mimetypes.best == 'application/json'
                       or request.headers.get('Sec-Fetch-Mode') in ('cors', 'no-cors')
                       or 'fetch' in request.headers.get('Sec-Fetch-Dest', ''))
            if is_ajax_check:
                return jsonify({"success": False, "error": "The email address you entered doesn't look right. Please check it and try again."}), 400
            flash("The email address you entered doesn't look right. Please check it and try again.", "warning")
            return redirect(url_for('bug_report'))
    
    include_diagnostics = "include_diagnostics" in request.form

    is_ajax = (request.headers.get('X-Requested-With') == 'XMLHttpRequest'
               or request.accept_mimetypes.best == 'application/json'
               or request.headers.get('Sec-Fetch-Mode') in ('cors', 'no-cors')
               or 'fetch' in request.headers.get('Sec-Fetch-Dest', ''))

    if not title or not description:
        if is_ajax:
            return jsonify({"success": False, "error": "Title and description are required."}), 400
        flash("Title and description are required.", "error")
        return redirect(url_for('bug_report'))

    diagnostics_data = None
    if include_diagnostics:
        diagnostics_data = json.dumps({
            "url": request.form.get("diagnostics_url", ""),
            "user_agent": request.form.get("diagnostics_ua", ""),
            "screen": request.form.get("diagnostics_screen", ""),
            "timestamp": request.form.get("diagnostics_time", "")
        })

    screenshot_filename = None
    screenshot = request.files.get("screenshot")
    if screenshot and screenshot.filename:
        from werkzeug.utils import secure_filename as wz_secure
        ext = wz_secure(screenshot.filename).rsplit(".", 1)[-1].lower()
        if ext in ("png", "jpg", "jpeg", "webp"):
            screenshot.seek(0, 2)
            file_size = screenshot.tell()
            screenshot.seek(0)
            if file_size <= 5 * 1024 * 1024:
                upload_dir = os.path.join(app.root_path, "static", "uploads")
                os.makedirs(upload_dir, exist_ok=True)
                screenshot_filename = f"{uuid.uuid4().hex[:12]}.{ext}"
                screenshot.save(os.path.join(upload_dir, screenshot_filename))

    username = None
    if current_user.is_authenticated:
        username = current_user.display_name

    report = BugReport(
        category=category,
        title=title[:80],
        description=description[:4000],
        contact_email=contact_email,
        include_diagnostics=include_diagnostics,
        diagnostics_data=diagnostics_data,
        screenshot_filename=screenshot_filename,
        username=username
    )
    db.session.add(report)
    db.session.commit()

    if is_ajax:
        response = jsonify({"success": True})
    else:
        flash("Feedback submitted successfully! Thank you.", "success")
        response = redirect(url_for('bug_report'))

    if contact_email:
        try:
            from email_responder import send_bug_report_reply
            send_bug_report_reply(
                to_email=contact_email,
                category=category,
                title=title,
                description=description
            )
        except Exception as email_err:
            app.logger.error(f"Bug report auto-reply email error: {email_err}")

    return response


@app.route("/login-redirect")
def login_and_redirect():
    """Redirect authenticated users to next_url or initiate Replit OAuth login."""
    next_url = request.args.get("next", "/")
    if current_user.is_authenticated:
        return redirect(next_url)
    session["next_url"] = request.url_root.rstrip("/") + next_url
    return redirect(url_for('replit_auth.login'))


@app.route("/support-chat")
def support_chat():
    """Render the AI-powered support chat page with session and CSRF management."""
    import uuid
    from datetime import datetime, timedelta
    chat_session_id = session.get('chat_session_id')
    if not chat_session_id:
        chat_session_id = uuid.uuid4().hex
        session['chat_session_id'] = chat_session_id

    if 'chat_csrf' not in session:
        session['chat_csrf'] = uuid.uuid4().hex

    history = SupportChat.query.filter_by(session_id=chat_session_id).order_by(SupportChat.created_at.asc()).all()

    WEEKLY_LIMIT = 50
    now = datetime.utcnow()
    week_start = now - timedelta(days=7)
    if current_user.is_authenticated:
        used = SupportChat.query.filter(
            SupportChat.user_id == current_user.id,
            SupportChat.role == 'user',
            SupportChat.created_at >= week_start
        ).count()
    else:
        used = SupportChat.query.filter(
            SupportChat.session_id == chat_session_id,
            SupportChat.role == 'user',
            SupportChat.created_at >= week_start
        ).count()

    remaining = max(0, WEEKLY_LIMIT - used)

    oldest_msg = None
    if used > 0:
        if current_user.is_authenticated:
            oldest_msg = SupportChat.query.filter(
                SupportChat.user_id == current_user.id,
                SupportChat.role == 'user',
                SupportChat.created_at >= week_start
            ).order_by(SupportChat.created_at.asc()).first()
        else:
            oldest_msg = SupportChat.query.filter(
                SupportChat.session_id == chat_session_id,
                SupportChat.role == 'user',
                SupportChat.created_at >= week_start
            ).order_by(SupportChat.created_at.asc()).first()

    renews_at = None
    if oldest_msg:
        renews_at = (oldest_msg.created_at + timedelta(days=7)).isoformat() + "Z"

    return render_template("support_chat.html", history=history, csrf_token=session['chat_csrf'],
                           remaining=remaining, weekly_limit=WEEKLY_LIMIT, renews_at=renews_at)


@app.route("/api/support-chat", methods=["POST"])
def api_support_chat():
    """Process a user chat message and return an AI-generated support response."""
    import uuid
    import os
    from datetime import datetime, timedelta
    from email_responder import generate_chat_reply
    from werkzeug.utils import secure_filename

    if request.content_type and 'multipart/form-data' in request.content_type:
        token = request.form.get("csrf_token", "")
        user_message = (request.form.get("message") or "").strip()
        attachment = request.files.get("attachment")
    else:
        data = request.get_json(silent=True) or {}
        token = data.get("csrf_token", "")
        user_message = (data.get("message") or "").strip()
        attachment = None

    if not token or token != session.get('chat_csrf'):
        return jsonify({"error": "Invalid request."}), 403

    if not user_message and not attachment:
        return jsonify({"error": "Please provide a message or attachment."}), 400
    if len(user_message) > 1000:
        return jsonify({"error": "Message too long (max 1000 characters)."}), 400

    MAX_FILE_SIZE = 10 * 1024 * 1024
    saved_filename = None
    original_name = None
    if attachment and attachment.filename:
        original_name = attachment.filename
        file_data = attachment.read()
        if len(file_data) > MAX_FILE_SIZE:
            return jsonify({"error": "File too large. Maximum size is 10 MB."}), 400
        attachment.seek(0)
        ext = os.path.splitext(secure_filename(original_name))[1].lower()
        saved_filename = uuid.uuid4().hex + ext
        upload_path = os.path.join("static", "chat_uploads", saved_filename)
        attachment.save(upload_path)

    chat_session_id = session.get('chat_session_id')
    if not chat_session_id:
        chat_session_id = uuid.uuid4().hex
        session['chat_session_id'] = chat_session_id

    username = None
    user_id = None
    if current_user.is_authenticated:
        username = current_user.display_name
        user_id = current_user.id

    WEEKLY_LIMIT = 50
    now = datetime.utcnow()
    week_start = now - timedelta(days=7)
    if current_user.is_authenticated:
        used = SupportChat.query.filter(
            SupportChat.user_id == current_user.id,
            SupportChat.role == 'user',
            SupportChat.created_at >= week_start
        ).count()
    else:
        used = SupportChat.query.filter(
            SupportChat.session_id == chat_session_id,
            SupportChat.role == 'user',
            SupportChat.created_at >= week_start
        ).count()

    if used >= WEEKLY_LIMIT:
        if saved_filename:
            try:
                os.remove(os.path.join("static", "chat_uploads", saved_filename))
            except OSError:
                pass
        oldest_msg = None
        if current_user.is_authenticated:
            oldest_msg = SupportChat.query.filter(
                SupportChat.user_id == current_user.id,
                SupportChat.role == 'user',
                SupportChat.created_at >= week_start
            ).order_by(SupportChat.created_at.asc()).first()
        else:
            oldest_msg = SupportChat.query.filter(
                SupportChat.session_id == chat_session_id,
                SupportChat.role == 'user',
                SupportChat.created_at >= week_start
            ).order_by(SupportChat.created_at.asc()).first()

        renews_at = None
        if oldest_msg:
            renews_at = (oldest_msg.created_at + timedelta(days=7)).isoformat() + "Z"

        return jsonify({"error": "You've reached the weekly limit of 50 messages. Your quota renews soon — check back later!",
                        "quota_exceeded": True, "remaining": 0, "renews_at": renews_at}), 429

    display_message = user_message if user_message else ""
    if saved_filename and not user_message:
        display_message = f"[Attached: {original_name}]"

    user_entry = SupportChat(
        session_id=chat_session_id,
        user_id=user_id,
        username=username,
        role="user",
        message=display_message,
        attachment_filename=saved_filename,
        attachment_original_name=original_name
    )
    db.session.add(user_entry)
    db.session.commit()

    history = SupportChat.query.filter_by(session_id=chat_session_id).order_by(SupportChat.created_at.asc()).all()
    conv = []
    for h in history:
        entry = {"role": h.role, "message": h.message}
        if h.attachment_filename and h.role == 'user':
            entry["attachment"] = {
                "path": os.path.join("static", "chat_uploads", h.attachment_filename),
                "original_name": h.attachment_original_name or h.attachment_filename,
                "filename": h.attachment_filename
            }
        conv.append(entry)

    ai_reply = generate_chat_reply(conv, user_name=username)
    if not ai_reply:
        ai_reply = "Sorry, I'm having trouble responding right now. Please try again in a moment!"

    ai_entry = SupportChat(
        session_id=chat_session_id,
        user_id=user_id,
        username="CricPredictor AI",
        role="assistant",
        message=ai_reply
    )
    db.session.add(ai_entry)
    db.session.commit()

    new_remaining = max(0, WEEKLY_LIMIT - (used + 1))
    renews_at = None
    oldest_msg = None
    if current_user.is_authenticated:
        oldest_msg = SupportChat.query.filter(
            SupportChat.user_id == current_user.id,
            SupportChat.role == 'user',
            SupportChat.created_at >= week_start
        ).order_by(SupportChat.created_at.asc()).first()
    else:
        oldest_msg = SupportChat.query.filter(
            SupportChat.session_id == chat_session_id,
            SupportChat.role == 'user',
            SupportChat.created_at >= week_start
        ).order_by(SupportChat.created_at.asc()).first()
    if oldest_msg:
        renews_at = (oldest_msg.created_at + timedelta(days=7)).isoformat() + "Z"

    attachment_info = None
    if saved_filename:
        attachment_info = {
            "url": "/static/chat_uploads/" + saved_filename,
            "name": original_name,
            "is_image": ext in ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.svg')
        }

    return jsonify({"reply": ai_reply, "remaining": new_remaining, "renews_at": renews_at, "attachment": attachment_info})


@app.route("/api/support-chat/clear", methods=["POST"])
def api_support_chat_clear():
    """Reset the chat session by generating a new session ID."""
    import uuid
    data = request.get_json(silent=True) or {}
    token = data.get("csrf_token", "")
    if not token or token != session.get('chat_csrf'):
        return jsonify({"error": "Invalid request."}), 403
    session['chat_session_id'] = uuid.uuid4().hex
    return jsonify({"success": True})
